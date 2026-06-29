#!/usr/bin/env python3
"""
monitor/inspect_r2_schema.py

Validates Excel files produced by the boshamlan.com scrapers against the
excel_schema defined in websites-config.yml.

Usage
-----
    # Daily validation (yesterday by default)
    python monitor/inspect_r2_schema.py

    # Custom date
    python monitor/inspect_r2_schema.py --date 2026-06-13

    # Look back 7 days and refresh persisted stats
    python monitor/inspect_r2_schema.py --days-lookback 7 --update-stats

    # Deep quality checks + exit 1 on any failure
    python monitor/inspect_r2_schema.py --quality --fail-on-error

Required environment variables (same as scraper workflows)
----------------------------------------------------------
    CF_R2_ACCESS_KEY_ID
    CF_R2_SECRET_ACCESS_KEY
    CF_R2_ENDPOINT_URL
    CF_R2_BUCKET_NAME

Config
------
    websites-config.yml is read from R2 at
    {r2_prefix}/monitor/websites-config.yml (default: boshamlan-data/monitor/).
    A local repo copy is used only as a development fallback.

Outputs
-------
    stdout              — summary table
    $GITHUB_STEP_SUMMARY — markdown summary (when running in GitHub Actions)
    R2: {r2_prefix}/monitor/{date}/report.json
    R2: {r2_prefix}/monitor/monitor_stats.yml  (only with --update-stats)
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from ads_counter import count_scraper_ads
from r2_file_counter import (
    category_slug_from_excel_pattern,
    count_scraper_r2_files,
    count_site_r2_files,
)

import boto3
import openpyxl
import yaml
from botocore.config import Config
from botocore.exceptions import ClientError

# ── Paths ─────────────────────────────────────────────────────────────────────
CONFIG_PATH = Path(__file__).parent.parent / "websites-config.yml"
CONFIG_R2_SUFFIX = "monitor/websites-config.yml"
DEFAULT_R2_PREFIX = "boshamlan-data"
EXCEL_FOLDER = "excel files"  # folder name used by both S3Uploader.py variants

# Adaptive row-count bounds kick in after this many observed runs in monitor_stats.yml.
ADAPTIVE_MIN_RUNS = 3
ROW_COUNT_MARGIN_ABS = 5
ROW_COUNT_MARGIN_RATIO = 0.25
DEFAULT_MIN_ROWS_FOR_QUALITY = 5


# ── CLI ───────────────────────────────────────────────────────────────────────
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Validate R2 Excel schema for boshamlan.com scrapers"
    )
    p.add_argument(
        "--date",
        default=None,
        metavar="YYYY-MM-DD",
        help="Date to check (default: yesterday UTC)",
    )
    p.add_argument(
        "--days-lookback",
        type=int,
        default=1,
        metavar="N",
        help="Number of consecutive days to check ending on --date",
    )
    p.add_argument(
        "--update-stats",
        action="store_true",
        help="Merge observations into monitor_stats.yml stored in R2",
    )
    p.add_argument(
        "--quality",
        action="store_true",
        help="Run deep data-quality checks via pandas",
    )
    p.add_argument(
        "--fail-on-error",
        action="store_true",
        help="Exit 1 if any check failed",
    )
    return p.parse_args()


# ── R2 client ─────────────────────────────────────────────────────────────────
def build_r2_client():
    return boto3.client(
        "s3",
        endpoint_url=os.environ["CF_R2_ENDPOINT_URL"],
        aws_access_key_id=os.environ["CF_R2_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["CF_R2_SECRET_ACCESS_KEY"],
        region_name="us-east-1",
        config=Config(
            signature_version="s3v4",
            s3={"addressing_style": "path"},
        ),
    )


# ── Helpers ───────────────────────────────────────────────────────────────────
def resolve_base_path(r2_path: str) -> str:
    """Strip the '{...}/' bucket placeholder from r2_path."""
    return re.sub(r"^\{[^}]+\}/", "", r2_path)


def date_partition(d: datetime) -> str:
    return f"year={d.year}/month={d.month:02d}/day={d.day:02d}"


def list_objects(client, bucket: str, prefix: str):
    """Yield all S3 object metadata dicts under *prefix* (handles pagination)."""
    paginator = client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            yield obj


def download_object(client, bucket: str, key: str) -> bytes:
    resp = client.get_object(Bucket=bucket, Key=key)
    return resp["Body"].read()


def load_existing_stats(client, bucket: str, stats_key: str) -> dict:
    try:
        data = download_object(client, bucket, stats_key)
        return yaml.safe_load(data) or {}
    except ClientError as exc:
        if exc.response["Error"]["Code"] in ("NoSuchKey", "404"):
            return {}
        raise


def load_config(client, bucket: str, r2_prefix: str) -> dict:
    """Load websites-config.yml from R2; fall back to local file for development."""
    config_key = f"{r2_prefix}/{CONFIG_R2_SUFFIX}"
    try:
        data = download_object(client, bucket, config_key)
        print(f"  Config loaded  ← r2://{bucket}/{config_key}")
        return yaml.safe_load(data) or {}
    except ClientError as exc:
        if exc.response["Error"]["Code"] not in ("NoSuchKey", "404"):
            raise
        if CONFIG_PATH.exists():
            print(f"  Config loaded  ← {CONFIG_PATH} (local fallback)")
            return yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")) or {}
        print(f"ERROR: websites-config.yml not found at r2://{bucket}/{config_key}")
        print(f"       and no local fallback at {CONFIG_PATH}")
        sys.exit(1)


def upload_bytes(
    client, bucket: str, key: str, data: bytes, content_type: str = "application/octet-stream"
):
    client.put_object(Bucket=bucket, Key=key, Body=data, ContentType=content_type)


def json_safe(value):
    """Recursively convert numpy/pandas scalars to native Python types for JSON.

    numpy scalars (.item() exists) are converted FIRST so that numpy.bool_,
    numpy.int64, numpy.float64, etc. are always reduced to their Python
    equivalents before the isinstance guards run — regardless of numpy version.
    """
    if hasattr(value, "item"):          # numpy / pandas scalar → Python native
        return json_safe(value.item())
    if isinstance(value, dict):
        return {k: json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [json_safe(v) for v in value]
    return value                         # str, int, float, bool, None — safe


# ── Validation ────────────────────────────────────────────────────────────────
def resolve_row_count_range(
    sheet_schema: dict,
    scraper_name: str,
    sheet_name: str,
    persisted_stats: dict | None,
) -> tuple[list[int], str]:
    """Return (effective [min, max], source label) for row_count validation."""
    static_range = list(sheet_schema.get("row_count_range", [0, 999_999]))
    if not persisted_stats:
        return static_range, "static"

    scraper_stats = persisted_stats.get(scraper_name, {})
    if scraper_stats.get("runs_observed", 0) < ADAPTIVE_MIN_RUNS:
        return static_range, "static"

    sheet_stats = scraper_stats.get("sheets", {}).get(sheet_name, {})
    hist_min = sheet_stats.get("row_count_min")
    hist_max = sheet_stats.get("row_count_max")
    if hist_min is None or hist_max is None:
        return static_range, "static"

    margin_low = max(ROW_COUNT_MARGIN_ABS, int(hist_min * ROW_COUNT_MARGIN_RATIO))
    margin_high = max(ROW_COUNT_MARGIN_ABS, int(hist_max * ROW_COUNT_MARGIN_RATIO))
    adaptive_low = max(0, hist_min - margin_low)
    adaptive_high = hist_max + margin_high
    effective = [
        min(static_range[0], adaptive_low),
        max(static_range[1], adaptive_high),
    ]
    return effective, "adaptive"


def validate_file(
    wb: openpyxl.Workbook,
    schema_entry: dict,
    file_size_bytes: int,
    scraper_name: str = "",
    persisted_stats: dict | None = None,
) -> list:
    """Return a list of check-result dicts for one Excel file."""
    checks = []
    sheet_names = wb.sheetnames
    min_kb = schema_entry.get("min_file_size_kb", 0)

    # File size
    actual_kb = file_size_bytes / 1024
    checks.append(
        {
            "check": "file_size",
            "pass": actual_kb >= min_kb,
            "detail": f"{actual_kb:.1f} KB (min {min_kb} KB)",
        }
    )

    for sheet_schema in schema_entry.get("sheets", []):
        sheet_name = sheet_schema["name"]
        required_cols = sheet_schema.get("required_columns", [])
        row_range, range_source = resolve_row_count_range(
            sheet_schema, scraper_name, sheet_name, persisted_stats
        )

        # Sheet exists?
        if sheet_name not in sheet_names:
            checks.append(
                {
                    "check": f"sheet_exists:{sheet_name}",
                    "pass": False,
                    "detail": f"Sheet '{sheet_name}' not found; available: {sheet_names}",
                }
            )
            continue
        checks.append({"check": f"sheet_exists:{sheet_name}", "pass": True, "detail": "ok"})

        ws = wb[sheet_name]

        # Column headers (first row)
        first_row = next(ws.iter_rows(min_row=1, max_row=1), [])
        actual_cols = [cell.value for cell in first_row if cell.value is not None]

        # Row count (data rows only, header excluded).
        # ws.max_row is None in openpyxl read-only mode when the sheet is
        # completely empty (e.g. offices with zero listings from yesterday).
        row_count = max(0, (ws.max_row or 1) - 1)

        optional_cols = sheet_schema.get("optional_columns", [])
        cols_to_check = list(required_cols)
        if row_count > 0:
            cols_to_check.extend(optional_cols)

        if row_count == 0 and not actual_cols:
            checks.append(
                {
                    "check": f"columns:{sheet_name}",
                    "pass": True,
                    "detail": "empty sheet (no headers)",
                }
            )
        else:
            missing = [c for c in cols_to_check if c not in actual_cols]
            checks.append(
                {
                    "check": f"columns:{sheet_name}",
                    "pass": not missing,
                    "detail": f"missing: {missing}" if missing else "ok",
                }
            )

        in_range = row_range[0] <= row_count <= row_range[1]
        range_label = "adaptive" if range_source == "adaptive" else "static"
        checks.append(
            {
                "check": f"row_count:{sheet_name}",
                "pass": in_range,
                "detail": (
                    f"{row_count} rows (expected [{row_range[0]}, {row_range[1]}], {range_label})"
                ),
            }
        )

    return checks


def _parse_published_date(value):
    """Parse date_published / Date Published from Excel (ISO, DD-MM-YYYY, datetime)."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "date") and callable(value.date):
        return value.date()
    text = str(value).strip()
    for fmt, length in (("%d-%m-%Y", 10), ("%Y-%m-%d", 10), ("%Y-%m-%d %H:%M:%S", 19)):
        try:
            return datetime.strptime(text[:length], fmt).date()
        except ValueError:
            continue
    return None


def _published_date_is_stale(value, check_date: datetime) -> bool:
    """True when published date falls outside the scraper's expected window.

    Property and office scrapers keep listings from yesterday 00:00 UTC through
    the partition date.  For a partition dated *check_date*, valid ads span
    check_date - 1 day through check_date inclusive (calendar date of
    date_published / Date Published).
    """
    published = _parse_published_date(value)
    if published is None:
        return False
    check_day = check_date.date() if isinstance(check_date, datetime) else check_date
    window_start = check_day - timedelta(days=1)
    return published < window_start or published > check_day


def quality_checks(wb: openpyxl.Workbook, schema_entry: dict, check_date: datetime) -> list:
    """Deep data-quality checks using pandas (only with --quality flag)."""
    import pandas as pd

    checks = []
    check_day = check_date.date() if isinstance(check_date, datetime) else check_date
    window_start = (check_day - timedelta(days=1)).strftime("%Y-%m-%d")
    window_end = check_day.strftime("%Y-%m-%d")

    for sheet_schema in schema_entry.get("sheets", []):
        sheet_name = sheet_schema["name"]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.values)
        if len(rows) < 2:
            continue
        df = pd.DataFrame(rows[1:], columns=rows[0])
        n = len(df)
        if n == 0:
            continue

        thresholds = sheet_schema.get("quality_thresholds", {})
        min_rows = thresholds.get("min_rows_for_quality", DEFAULT_MIN_ROWS_FOR_QUALITY)
        if n < min_rows:
            continue

        # Null title / Name %
        title_col = next((c for c in ["title", "Name"] if c in df.columns), None)
        if title_col:
            pct = df[title_col].isna().sum() / n * 100
            checks.append(
                {
                    "check": f"quality:null_{title_col}_pct:{sheet_name}",
                    "pass": pct < 10.0,
                    "detail": f"{pct:.1f}%",
                }
            )

        # Null price / Price %.
        # Threshold is configurable via quality_thresholds.null_price_max_pct in
        # the sheet schema (default 50%).  Offices routinely omit prices so their
        # main sheet sets this to 100 to suppress false-positive failures.
        price_col = next((c for c in ["price", "Price"] if c in df.columns), None)
        if price_col:
            price_threshold = thresholds.get("null_price_max_pct", 50.0)
            if price_threshold < 100:
                pct = df[price_col].isna().sum() / n * 100
                checks.append(
                    {
                        "check": f"quality:null_{price_col}_pct:{sheet_name}",
                        "pass": pct <= price_threshold,
                        "detail": f"{pct:.1f}% (max {price_threshold:.0f}%)",
                    }
                )

        # Stale date % — properties (ISO date_published) and offices (Date Published)
        # both use the scraper's 2-day window: partition_date-1 .. partition_date.
        date_col = next(
            (c for c in ["date_published", "Date Published"] if c in df.columns), None
        )
        if date_col:
            stale = df[date_col].apply(lambda v: _published_date_is_stale(v, check_date))
            pct = stale.sum() / n * 100
            stale_threshold = thresholds.get("stale_date_max_pct", 20.0)
            checks.append(
                {
                    "check": f"quality:stale_date_pct:{sheet_name}",
                    "pass": pct <= stale_threshold,
                    "detail": (
                        f"{pct:.1f}% stale "
                        f"(max {stale_threshold:.0f}%; window {window_start}–{window_end})"
                    ),
                }
            )

        # Duplicate link / URL count
        link_col = next((c for c in ["link", "URL"] if c in df.columns), None)
        if link_col:
            dups = int(df[link_col].dropna().duplicated().sum())
            checks.append(
                {
                    "check": f"quality:duplicate_{link_col}:{sheet_name}",
                    "pass": dups == 0,
                    "detail": f"{dups} duplicate(s)",
                }
            )

    return checks


# ── Stats merging ─────────────────────────────────────────────────────────────
def merge_stats(existing: dict, observations: dict) -> dict:
    """
    Merge new per-scraper observations into the persisted stats dict.
    All values kept in R2 — never written locally or committed to git.
    """
    stats = {k: v for k, v in existing.items()}  # shallow copy at top level

    for scraper_name, obs in observations.items():
        s = stats.setdefault(
            scraper_name,
            {"runs_observed": 0, "files_found_min": None, "files_found_max": None, "sheets": {}},
        )
        s["runs_observed"] = s.get("runs_observed", 0) + 1

        ff = obs.get("files_found", 0)
        if s["files_found_min"] is None or ff < s["files_found_min"]:
            s["files_found_min"] = ff
        if s["files_found_max"] is None or ff > s["files_found_max"]:
            s["files_found_max"] = ff

        for sheet_name, sheet_obs in obs.get("sheets", {}).items():
            sh = s["sheets"].setdefault(
                sheet_name, {"row_count_min": None, "row_count_max": None, "col_union": []}
            )
            rc = sheet_obs.get("row_count", 0)
            if sh["row_count_min"] is None or rc < sh["row_count_min"]:
                sh["row_count_min"] = rc
            if sh["row_count_max"] is None or rc > sh["row_count_max"]:
                sh["row_count_max"] = rc
            existing_cols = set(sh.get("col_union", []))
            existing_cols.update(sheet_obs.get("columns", []))
            sh["col_union"] = sorted(existing_cols)

    return stats


# ── Reporting ─────────────────────────────────────────────────────────────────
def print_summary(report: dict):
    print("\n" + "=" * 72)
    print("  R2 SCHEMA MONITOR — boshamlan.com")
    print("=" * 72)
    print(f"  Date checked : {report['date']}")
    print(f"  Days lookback: {report['days_lookback']}")
    print(f"  Overall      : {'✓ PASS' if report['overall_pass'] else '✗ FAIL'}")
    print("=" * 72)
    print(
        f"  {'Scraper':<32} {'Files':>5}  {'R2':>7}  {'Ads':>6}  "
        f"{'Checks':>12}  {'Status':>6}"
    )
    print("  " + "-" * 84)
    for entry in report["scrapers"]:
        status = "PASS" if entry["all_passed"] else "FAIL"
        print(
            f"  {entry['scraper']:<32} {entry['files_found']:>5}  "
            f"{entry.get('r2_file_count', 0):>7}  "
            f"{entry.get('unique_ads', 0):>6}  "
            f"{entry['checks_passed']:>5}/{entry['checks_total']:<5}  {status:>6}"
        )
    if "total_unique_ads" in report or "total_r2_files" in report:
        print("  " + "-" * 84)
        if "total_unique_ads" in report:
            print(f"  {'TOTAL UNIQUE ADS':<32} {'':>5}  {'':>7}  {report['total_unique_ads']:>6}")
        if "total_r2_files" in report:
            print(f"  {'TOTAL R2 FILES':<32} {'':>5}  {report['total_r2_files']:>7}")
    print("=" * 72 + "\n")


def write_step_summary(report: dict):
    summary_path = os.environ.get("GITHUB_STEP_SUMMARY")
    if not summary_path:
        return
    with open(summary_path, "a", encoding="utf-8") as fh:
        fh.write("## R2 Schema Monitor — boshamlan.com\n\n")
        fh.write(f"**Date:** `{report['date']}`  \n")
        fh.write(f"**Overall:** {'✅ PASS' if report['overall_pass'] else '❌ FAIL'}\n\n")
        fh.write("| Scraper | Files | R2 files | Unique ads | Checks | Status |\n")
        fh.write("|---------|------:|---------:|-----------:|-------:|:------:|\n")
        for entry in report["scrapers"]:
            icon = "✅" if entry["all_passed"] else "❌"
            fh.write(
                f"| {entry['scraper']} | {entry['files_found']} | "
                f"{entry.get('r2_file_count', 0)} | "
                f"{entry.get('unique_ads', 0)} | "
                f"{entry['checks_passed']}/{entry['checks_total']} | {icon} |\n"
            )
        if "total_unique_ads" in report or "total_r2_files" in report:
            ads_total = report.get("total_unique_ads", "")
            r2_total = report.get("total_r2_files", "")
            fh.write(f"| **Total** | | **{r2_total}** | **{ads_total}** | | |\n")
        fh.write("\n")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    args = parse_args()

    # Bootstrap R2 client before config (config lives in R2, not in the repo)
    r2_prefix = os.environ.get("MONITOR_R2_PREFIX", DEFAULT_R2_PREFIX)
    bucket = os.environ.get("CF_R2_BUCKET_NAME")
    if not bucket:
        print("ERROR: CF_R2_BUCKET_NAME environment variable is required.")
        sys.exit(1)
    client = build_r2_client()

    config = load_config(client, bucket, r2_prefix)
    meta = config.get("meta", {})
    r2_prefix = meta.get("r2_prefix", r2_prefix)

    scrapers_conf = {s["name"]: s for s in config.get("scrapers", [])}
    schema_map = {e["scraper"]: e for e in config.get("excel_schema", [])}

    if not schema_map:
        print("ERROR: No excel_schema found in websites-config.yml — run the excel-schema prompt first.")
        sys.exit(1)

    stats_key = f"{r2_prefix}/monitor/monitor_stats.yml"
    persisted_stats = load_existing_stats(client, bucket, stats_key)
    if persisted_stats:
        print(f"  Stats loaded   ← r2://{bucket}/{stats_key}")

    # Date range
    if args.date:
        end_date = datetime.strptime(args.date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    else:
        end_date = datetime.now(timezone.utc) - timedelta(days=1)

    dates = [end_date - timedelta(days=i) for i in range(args.days_lookback)]

    all_results: list = []
    new_stats_obs: dict = {}
    any_failure = False
    scraper_r2_counts: dict[tuple[str, str | None], int] = {}

    for check_date in dates:
        date_str = check_date.strftime("%Y-%m-%d")
        print(f"\n{'='*72}")
        print(f"  Date: {date_str}")
        print(f"{'='*72}")

        for scraper_name, schema_entry in schema_map.items():
            scraper_conf = scrapers_conf.get(scraper_name, {})
            r2_path_raw = scraper_conf.get("r2_path", "")
            base_path = resolve_base_path(r2_path_raw)
            pattern = schema_entry.get("excel_file_pattern", "")
            category_slug = category_slug_from_excel_pattern(pattern)
            r2_count_key = (base_path, category_slug)
            if r2_count_key not in scraper_r2_counts:
                if category_slug:
                    print(
                        f"    r2 inventory: counting {category_slug} under {base_path}/ ..."
                    )
                else:
                    print(f"    r2 inventory: counting objects under {base_path}/ ...")
                scraper_r2_counts[r2_count_key] = count_scraper_r2_files(
                    client, bucket, base_path, category_slug
                )
                print(f"    r2_file_count: {scraper_r2_counts[r2_count_key]:,}")
            partition = date_partition(check_date)
            prefix = f"{base_path}/{partition}/{EXCEL_FOLDER}/"
            is_wildcard = "{" in pattern

            print(f"\n  [{scraper_name}]")
            print(f"    prefix : {prefix}")

            # List objects
            all_objects = list(list_objects(client, bucket, prefix))
            xlsx_objects = [o for o in all_objects if o["Key"].endswith(".xlsx")]

            if is_wildcard:
                target_objects = xlsx_objects
            else:
                target_objects = [
                    o for o in xlsx_objects if o["Key"].split("/")[-1] == pattern
                ]

            files_found = len(target_objects)
            print(f"    files  : {files_found} found")

            scraper_result = {
                "scraper": scraper_name,
                "date": date_str,
                "files_found": files_found,
                "r2_file_count": scraper_r2_counts[r2_count_key],
                "checks_passed": 0,
                "checks_total": 0,
                "all_passed": True,
                "files": [],
            }
            obs_sheets: dict = {}
            excel_downloads: list[tuple[str, bytes]] = []

            for obj in target_objects:
                key = obj["Key"]
                file_size = obj["Size"]
                filename = key.split("/")[-1]
                print(f"    checking: {filename} ({file_size:,} bytes)")

                try:
                    raw = download_object(client, bucket, key)
                    excel_downloads.append((key, raw))
                    wb = openpyxl.load_workbook(BytesIO(raw), read_only=True)

                    checks = validate_file(
                        wb,
                        schema_entry,
                        file_size,
                        scraper_name=scraper_name,
                        persisted_stats=persisted_stats,
                    )
                    if args.quality:
                        checks += quality_checks(wb, schema_entry, check_date)

                    passed = sum(1 for c in checks if c["pass"])
                    total = len(checks)
                    file_ok = passed == total

                    if not file_ok:
                        any_failure = True
                        scraper_result["all_passed"] = False
                        for c in checks:
                            if not c["pass"]:
                                print(f"      ✗ {c['check']}: {c['detail']}")

                    scraper_result["checks_passed"] += passed
                    scraper_result["checks_total"] += total
                    scraper_result["files"].append(
                        {
                            "key": key,
                            "size_bytes": file_size,
                            "sheets": wb.sheetnames,
                            "checks": checks,
                            "all_passed": file_ok,
                        }
                    )

                    # Collect stats observations per sheet
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        row_count = max(0, (ws.max_row or 1) - 1)
                        first_row_iter = ws.iter_rows(min_row=1, max_row=1)
                        cols = [c.value for c in next(first_row_iter, []) if c.value]
                        if sheet_name not in obs_sheets:
                            obs_sheets[sheet_name] = {"row_count": row_count, "columns": cols}
                        else:
                            obs_sheets[sheet_name]["row_count"] = max(
                                obs_sheets[sheet_name]["row_count"], row_count
                            )

                    wb.close()

                except Exception as exc:
                    print(f"      ERROR reading {filename}: {exc}")
                    any_failure = True
                    scraper_result["all_passed"] = False
                    scraper_result["files"].append(
                        {"key": key, "size_bytes": file_size, "error": str(exc), "all_passed": False}
                    )

            # No files found — count as a failure for properties; warn for offices
            if files_found == 0:
                any_failure = True
                scraper_result["all_passed"] = False
                print(f"    WARNING: no files found for '{scraper_name}' on {date_str}")

            ads_stats = count_scraper_ads(
                client, bucket, base_path, check_date, excel_downloads
            )
            scraper_result["unique_ads"] = ads_stats.get("unique_ads") or 0
            scraper_result["total_rows"] = ads_stats.get("total_rows") or 0
            scraper_result["ads_source"] = ads_stats.get("ads_source", "none")
            print(
                f"    ads    : {scraper_result['unique_ads']} unique "
                f"({scraper_result['ads_source']})"
            )

            status_str = "PASS" if scraper_result["all_passed"] else "FAIL"
            print(
                f"    result : {status_str} "
                f"({scraper_result['checks_passed']}/{scraper_result['checks_total']} checks)"
            )

            all_results.append(scraper_result)
            new_stats_obs[scraper_name] = {"files_found": files_found, "sheets": obs_sheets}

    # Build report
    report_date = end_date.strftime("%Y-%m-%d")
    site_r2_prefix = meta.get("r2_prefix", r2_prefix).strip("/")
    if site_r2_prefix:
        print(f"\n  Site R2 inventory: counting objects under {site_r2_prefix}/ ...")
        total_r2_files = count_site_r2_files(client, bucket, site_r2_prefix)
        print(f"  total_r2_files: {total_r2_files:,}")
    else:
        report_scrapers = [r for r in all_results if r.get("date") == report_date]
        total_r2_files = sum(r.get("r2_file_count") or 0 for r in report_scrapers)

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "date": report_date,
        "days_lookback": args.days_lookback,
        "overall_pass": not any_failure,
        "total_unique_ads": sum(
            r.get("unique_ads") or 0
            for r in all_results
            if r.get("date") == report_date
        ),
        "total_r2_files": total_r2_files,
        "scrapers": all_results,
    }

    print_summary(report)
    write_step_summary(report)

    # Upload report to R2 (never written to local filesystem or git)
    report_key = f"{r2_prefix}/monitor/{end_date.strftime('%Y-%m-%d')}/report.json"
    upload_bytes(
        client, bucket, report_key,
        json.dumps(json_safe(report), ensure_ascii=False, indent=2).encode("utf-8"),
        "application/json",
    )
    print(f"  Report uploaded → r2://{bucket}/{report_key}")

    # Update persisted stats (only with --update-stats)
    if args.update_stats:
        updated = merge_stats(persisted_stats, new_stats_obs)
        upload_bytes(
            client, bucket, stats_key,
            yaml.dump(updated, allow_unicode=True, default_flow_style=False).encode("utf-8"),
            "application/x-yaml",
        )
        print(f"  Stats updated  → r2://{bucket}/{stats_key}")

    if args.fail_on_error and any_failure:
        sys.exit(1)


if __name__ == "__main__":
    main()
